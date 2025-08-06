from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from datetime import datetime, timedelta
from typing import List, Optional

from app.database import get_db
from app.models.user import User
from app.models.sensor_data import SensorData, SensorMapping
from app.schemas.sensor_data import SensorDataResponse, SensorDataStats, SensorDataPaginated
from app.utils.dependencies import get_current_user

router = APIRouter()

@router.get("/my-sensors")
async def get_my_sensors(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """自分に紐づくセンサー一覧取得"""
    sensors = db.query(SensorMapping)\
                .filter_by(user_id=current_user.user_id, is_active=True)\
                .all()
    
    return [
        {
            "sensor_id": sensor.sensor_id,
            "device_type": sensor.device_type,
            "subject_name": sensor.subject_name,
            "created_at": sensor.created_at
        }
        for sensor in sensors
    ]

@router.get("/my-data/stats")
async def get_my_data_stats(
    sensor_id: Optional[str] = Query(None, description="特定センサーID（未指定時は全センサー）"),
    start_date: Optional[datetime] = Query(None, description="開始日時"),
    end_date: Optional[datetime] = Query(None, description="終了日時"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """自分のセンサデータ統計取得"""
    
    # 基本クエリ
    query = db.query(SensorData).filter_by(user_id=current_user.user_id)
    
    # フィルタ条件追加
    if sensor_id:
        query = query.filter_by(sensor_id=sensor_id)
    
    if start_date:
        query = query.filter(SensorData.timestamp >= start_date)
    
    if end_date:
        query = query.filter(SensorData.timestamp <= end_date)
    
    # 統計計算
    stats = query.with_entities(
        func.count(SensorData.id).label('total_records'),
        func.min(SensorData.temperature).label('min_temperature'),
        func.max(SensorData.temperature).label('max_temperature'),
        func.avg(SensorData.temperature).label('avg_temperature'),
        func.min(SensorData.timestamp).label('start_time'),
        func.max(SensorData.timestamp).label('end_time')
    ).first()
    
    if stats.total_records == 0:
        return {
            "total_records": 0,
            "min_temperature": None,
            "max_temperature": None,
            "avg_temperature": None,
            "start_time": None,
            "end_time": None,
            "message": "No data found for the specified criteria"
        }
    
    return {
        "total_records": stats.total_records,
        "min_temperature": round(stats.min_temperature, 2),
        "max_temperature": round(stats.max_temperature, 2),
        "avg_temperature": round(stats.avg_temperature, 2),
        "start_time": stats.start_time,
        "end_time": stats.end_time
    }

@router.get("/my-data", response_model=SensorDataPaginated)
async def get_my_sensor_data(
    sensor_id: Optional[str] = Query(None, description="特定センサーID"),
    start_date: Optional[datetime] = Query(None, description="開始日時"),
    end_date: Optional[datetime] = Query(None, description="終了日時"),
    page: int = Query(0, ge=0, description="ページ番号（0から開始）"),
    limit: int = Query(100, ge=1, le=1000, description="1ページあたりの件数"),
    order: str = Query("desc", regex="^(asc|desc)$", description="並び順（asc/desc）"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """自分のセンサデータ取得（ページネーション対応）"""
    
    # 基本クエリ
    query = db.query(SensorData).filter_by(user_id=current_user.user_id)
    
    # フィルタ条件追加
    if sensor_id:
        query = query.filter_by(sensor_id=sensor_id)
    
    if start_date:
        query = query.filter(SensorData.timestamp >= start_date)
    
    if end_date:
        query = query.filter(SensorData.timestamp <= end_date)
    
    # 総件数取得
    total = query.count()
    
    # ソート・ページネーション
    if order == "desc":
        query = query.order_by(SensorData.timestamp.desc())
    else:
        query = query.order_by(SensorData.timestamp.asc())
    
    data = query.offset(page * limit).limit(limit).all()
    
    return SensorDataPaginated(
        data=data,
        total=total,
        page=page,
        limit=limit,
        has_next=(page + 1) * limit < total
    )

@router.get("/my-data/chart")
async def get_chart_data(
    sensor_id: Optional[str] = Query(None, description="特定センサーID"),
    start_date: Optional[datetime] = Query(None, description="開始日時"),
    end_date: Optional[datetime] = Query(None, description="終了日時"),
    interval: str = Query("1hour", regex="^(1min|5min|15min|1hour|1day)$", description="データ間隔"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """グラフ表示用のセンサデータ取得（間引き・集約済み）"""
    
    # デフォルトの日時範囲設定（指定がない場合は直近24時間）
    if not end_date:
        end_date = datetime.utcnow()
    if not start_date:
        start_date = end_date - timedelta(hours=24)
    
    # 基本クエリ
    query = db.query(SensorData).filter_by(user_id=current_user.user_id)
    
    if sensor_id:
        query = query.filter_by(sensor_id=sensor_id)
    
    query = query.filter(
        and_(
            SensorData.timestamp >= start_date,
            SensorData.timestamp <= end_date
        )
    )
    
    # データ間隔による集約
    if interval == "1min":
        # 1分間隔：そのまま取得（最大1440件/日）
        data = query.order_by(SensorData.timestamp).all()
    elif interval == "5min":
        # 5分間隔：5分ごとに平均値算出
        data = query.order_by(SensorData.timestamp).all()
        data = _aggregate_data(data, minutes=5)
    elif interval == "15min":
        # 15分間隔
        data = query.order_by(SensorData.timestamp).all()
        data = _aggregate_data(data, minutes=15)
    elif interval == "1hour":
        # 1時間間隔
        data = query.order_by(SensorData.timestamp).all()
        data = _aggregate_data(data, hours=1)
    elif interval == "1day":
        # 1日間隔
        data = query.order_by(SensorData.timestamp).all()
        data = _aggregate_data(data, days=1)
    
    # グラフ用フォーマットに変換
    chart_data = [
        {
            "timestamp": item["timestamp"].isoformat(),
            "temperature": round(item["temperature"], 2),
            "sensor_id": item["sensor_id"] if isinstance(item, dict) else item.sensor_id
        }
        for item in data[:1000]  # 最大1000ポイントに制限
    ]
    
    return {
        "data": chart_data,
        "total_points": len(chart_data),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "interval": interval
    }

def _aggregate_data(data, minutes=None, hours=None, days=None):
    """データ集約関数"""
    if not data:
        return []
    
    # 集約間隔計算
    if minutes:
        delta = timedelta(minutes=minutes)
    elif hours:
        delta = timedelta(hours=hours)
    elif days:
        delta = timedelta(days=days)
    else:
        return data
    
    aggregated = []
    current_group = []
    current_timestamp = None
    
    for item in data:
        # 集約時間窓の開始時刻計算
        item_time = item.timestamp
        window_start = item_time.replace(second=0, microsecond=0)
        
        if minutes:
            window_start = window_start.replace(minute=(window_start.minute // minutes) * minutes)
        elif hours:
            window_start = window_start.replace(minute=0)
        elif days:
            window_start = window_start.replace(hour=0, minute=0)
        
        # 新しい時間窓の場合
        if current_timestamp != window_start:
            # 前のグループを集約
            if current_group:
                avg_temp = sum(d.temperature for d in current_group) / len(current_group)
                aggregated.append({
                    "timestamp": current_timestamp,
                    "temperature": avg_temp,
                    "sensor_id": current_group[0].sensor_id,
                    "count": len(current_group)
                })
            
            # 新しいグループ開始
            current_timestamp = window_start
            current_group = [item]
        else:
            current_group.append(item)
    
    # 最後のグループを集約
    if current_group:
        avg_temp = sum(d.temperature for d in current_group) / len(current_group)
        aggregated.append({
            "timestamp": current_timestamp,
            "temperature": avg_temp,
            "sensor_id": current_group[0].sensor_id,
            "count": len(current_group)
        })
    
    return aggregated

@router.get("/my-data/export")
async def export_my_data(
    format: str = Query("csv", regex="^(csv|json)$", description="エクスポート形式"),
    sensor_id: Optional[str] = Query(None, description="特定センサーID"),
    start_date: Optional[datetime] = Query(None, description="開始日時"),
    end_date: Optional[datetime] = Query(None, description="終了日時"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """センサデータエクスポート"""
    from fastapi.responses import StreamingResponse
    import csv
    import json
    from io import StringIO
    
    # データ取得
    query = db.query(SensorData).filter_by(user_id=current_user.user_id)
    
    if sensor_id:
        query = query.filter_by(sensor_id=sensor_id)
    
    if start_date:
        query = query.filter(SensorData.timestamp >= start_date)
    
    if end_date:
        query = query.filter(SensorData.timestamp <= end_date)
    
    data = query.order_by(SensorData.timestamp).limit(10000).all()  # 最大10000件
    
    if format == "csv":
        # CSV形式
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['timestamp', 'sensor_id', 'temperature'])
        
        for item in data:
            writer.writerow([
                item.timestamp.isoformat(),
                item.sensor_id,
                item.temperature
            ])
        
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=sensor_data_{current_user.user_id}.csv"}
        )
    
    elif format == "json":
        # JSON形式
        json_data = [
            {
                "timestamp": item.timestamp.isoformat(),
                "sensor_id": item.sensor_id,
                "temperature": item.temperature
            }
            for item in data
        ]
        
        return StreamingResponse(
            iter([json.dumps(json_data, indent=2)]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=sensor_data_{current_user.user_id}.json"}
        )

@router.get("/my-data/export-enhanced")
async def export_data_enhanced(
    format: str = Query(..., regex="^(csv|json|excel)$", description="エクスポート形式"),
    sensor_id: Optional[str] = Query(None, description="特定センサーID"),
    start_date: Optional[datetime] = Query(None, description="開始日時"),
    end_date: Optional[datetime] = Query(None, description="終了日時"),
    min_temperature: Optional[float] = Query(None, description="最低温度"),
    max_temperature: Optional[float] = Query(None, description="最高温度"),
    limit: Optional[int] = Query(10000, description="最大レコード数"),
    skip: Optional[int] = Query(0, description="スキップするレコード数"),
    include_metadata: bool = Query(True, description="メタデータを含める"),
    include_headers: bool = Query(True, description="ヘッダーを含める"),
    timezone: str = Query("Asia/Tokyo", description="タイムゾーン"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """強化されたデータエクスポート機能"""
    from fastapi.responses import StreamingResponse
    import csv
    import json
    from io import StringIO, BytesIO
    import pytz
    from datetime import datetime as dt
    
    try:
        # データ取得
        query = db.query(SensorData).filter_by(user_id=current_user.user_id)
        
        # フィルター適用
        if sensor_id:
            query = query.filter_by(sensor_id=sensor_id)
        
        if start_date:
            query = query.filter(SensorData.timestamp >= start_date)
        
        if end_date:
            query = query.filter(SensorData.timestamp <= end_date)
            
        if min_temperature is not None:
            query = query.filter(SensorData.temperature >= min_temperature)
            
        if max_temperature is not None:
            query = query.filter(SensorData.temperature <= max_temperature)
        
        # レコード数制限とスキップ
        total_count = query.count()
        data = query.order_by(SensorData.timestamp.desc())\
                   .offset(skip)\
                   .limit(limit)\
                   .all()
        
        # タイムゾーン設定
        target_tz = pytz.timezone(timezone)
        
        # メタデータ準備
        metadata = {}
        if include_metadata:
            metadata = {
                'exported_at': dt.utcnow().replace(tzinfo=pytz.UTC).astimezone(target_tz).isoformat(),
                'exported_by': current_user.username,
                'total_records': len(data),
                'total_available': total_count,
                'filters': {
                    'sensor_id': sensor_id,
                    'start_date': start_date.isoformat() if start_date else None,
                    'end_date': end_date.isoformat() if end_date else None,
                    'min_temperature': min_temperature,
                    'max_temperature': max_temperature,
                },
                'timezone': timezone,
                'format': format
            }
        
        # フォーマット別処理
        if format == 'csv':
            return _export_csv_enhanced(data, metadata, include_headers, include_metadata, target_tz)
        elif format == 'json':
            return _export_json_enhanced(data, metadata, include_metadata, target_tz)
        elif format == 'excel':
            return _export_excel_enhanced(data, metadata, include_headers, include_metadata, target_tz)
            
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Export failed: {str(e)}"
        )

def _export_csv_enhanced(data, metadata, include_headers, include_metadata, target_tz):
    """CSV形式でのエクスポート（強化版）"""
    from fastapi.responses import StreamingResponse
    import csv
    from io import StringIO
    
    output = StringIO()
    
    # メタデータをコメントとして追加
    if include_metadata:
        output.write(f"# エクスポート日時: {metadata['exported_at']}\n")
        output.write(f"# エクスポートユーザー: {metadata['exported_by']}\n")
        output.write(f"# 総レコード数: {metadata['total_records']}\n")
        output.write(f"# タイムゾーン: {metadata['timezone']}\n")
        if metadata['filters']['sensor_id']:
            output.write(f"# センサーID: {metadata['filters']['sensor_id']}\n")
        if metadata['filters']['start_date']:
            output.write(f"# 開始日時: {metadata['filters']['start_date']}\n")
        if metadata['filters']['end_date']:
            output.write(f"# 終了日時: {metadata['filters']['end_date']}\n")
        output.write("#\n")
    
    writer = csv.writer(output)
    
    # ヘッダー
    if include_headers:
        headers = ['timestamp', 'sensor_id', 'temperature', 'local_time']
        if include_metadata:
            headers.extend(['created_at', 'data_source'])
        writer.writerow(headers)
    
    # データ行
    for item in data:
        local_time = item.timestamp.replace(tzinfo=pytz.UTC).astimezone(target_tz)
        
        row = [
            item.timestamp.isoformat(),
            item.sensor_id,
            item.temperature,
            local_time.strftime('%Y-%m-%d %H:%M:%S %Z')
        ]
        
        if include_metadata:
            row.extend([
                item.created_at.isoformat() if item.created_at else '',
                item.data_source or ''
            ])
        
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=sensor_data_enhanced.csv"}
    )

def _export_json_enhanced(data, metadata, include_metadata, target_tz):
    """JSON形式でのエクスポート（強化版）"""
    from fastapi.responses import StreamingResponse
    import json
    
    # データ変換
    json_data = []
    for item in data:
        local_time = item.timestamp.replace(tzinfo=pytz.UTC).astimezone(target_tz)
        
        record = {
            "timestamp": item.timestamp.isoformat(),
            "sensor_id": item.sensor_id,
            "temperature": item.temperature,
            "local_time": local_time.isoformat(),
        }
        
        if include_metadata:
            record.update({
                "created_at": item.created_at.isoformat() if item.created_at else None,
                "data_source": item.data_source,
                "id": item.id
            })
        
        json_data.append(record)
    
    # 最終JSON構造
    result = {
        "data": json_data
    }
    
    if include_metadata:
        result["metadata"] = metadata
    
    return StreamingResponse(
        iter([json.dumps(result, indent=2, ensure_ascii=False)]),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=sensor_data_enhanced.json"}
    )

def _export_excel_enhanced(data, metadata, include_headers, include_metadata, target_tz):
    """Excel形式でのエクスポート（強化版）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        # DataFrameを作成
        df_data = []
        for item in data:
            local_time = item.timestamp.replace(tzinfo=pytz.UTC).astimezone(target_tz)
            
            record = {
                'timestamp': item.timestamp,
                'sensor_id': item.sensor_id,
                'temperature': item.temperature,
                'local_time': local_time.strftime('%Y-%m-%d %H:%M:%S'),
            }
            
            if include_metadata:
                record.update({
                    'created_at': item.created_at,
                    'data_source': item.data_source or ''
                })
            
            df_data.append(record)
        
        df = pd.DataFrame(df_data)
        
        # Excelワークブック作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sensor Data"
        
        # スタイル定義
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        row_idx = 1
        
        # メタデータ情報
        if include_metadata:
            ws[f'A{row_idx}'] = "トライアスロンセンサデータエクスポート"
            ws[f'A{row_idx}'].font = Font(bold=True, size=14)
            row_idx += 2
            
            info_items = [
                ('エクスポート日時', metadata['exported_at']),
                ('エクスポートユーザー', metadata['exported_by']),
                ('総レコード数', metadata['total_records']),
                ('タイムゾーン', metadata['timezone']),
            ]
            
            if metadata['filters']['sensor_id']:
                info_items.append(('センサーID', metadata['filters']['sensor_id']))
            if metadata['filters']['start_date']:
                info_items.append(('開始日時', metadata['filters']['start_date']))
            if metadata['filters']['end_date']:
                info_items.append(('終了日時', metadata['filters']['end_date']))
            
            for label, value in info_items:
                ws[f'A{row_idx}'] = label
                ws[f'B{row_idx}'] = str(value)
                ws[f'A{row_idx}'].font = Font(bold=True)
                row_idx += 1
            
            row_idx += 2  # 空行
        
        # データテーブル
        if include_headers:
            headers = ['日時', 'センサーID', '温度(°C)', '現地時刻']
            if include_metadata:
                headers.extend(['作成日時', 'データソース'])
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            row_idx += 1
        
        # データ行
        for item in data:
            local_time = item.timestamp.replace(tzinfo=pytz.UTC).astimezone(target_tz)
            
            row_data = [
                item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                item.sensor_id,
                item.temperature,
                local_time.strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            if include_metadata:
                row_data.extend([
                    item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
                    item.data_source or ''
                ])
            
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            
            row_idx += 1
        
        # 列幅自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # ファイル出力
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=sensor_data_enhanced.xlsx"}
        )
        
    except ImportError:
        # openpyxlが利用できない場合はCSV形式にフォールバック
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export requires openpyxl package. Please use CSV or JSON format."
        )